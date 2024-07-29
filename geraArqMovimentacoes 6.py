import mysql.connector
from datetime import datetime, timedelta
import openpyxl
import xlrd
import re

# Funções
def converte_xls_xlsx(arq_xls, arq_xlsx):
  xls_book = xlrd.open_workbook(arq_xls)
  xlsx_book = openpyxl.Workbook()
  for sheet_index in range(xls_book.nsheets):
    xls_sheet = xls_book.sheet_by_index(sheet_index)
    if sheet_index == 0:
      xlsx_sheet = xlsx_book.active
      xlsx_sheet.title = xls_sheet.name
    else:
      xlsx_sheet = xlsx_book.create_sheet(title=xls_sheet.name)
    
    for row in range(xls_sheet.nrows):
      for col in range(xls_sheet.ncols):
        cell_value = xls_sheet.cell_value(row, col)
        xlsx_sheet.cell(row=row + 1, column=col + 1, value=cell_value)
  xlsx_book.save(arq_xlsx)
def verifica_abreviacao(string):
  regex = r'[.]'
  if re.search(regex, string) != None:
    return True
  else:
    return False
def limitar_string(string, limit):
  if string is not None:
    if len(string) > limit:
      return string[:limit]
  return string
def formatar_data(data):
  formatos_possiveis = ["%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y", "%d-%m-%Y", "%Y/%m/%d"]
  for formato in formatos_possiveis:
    try:
      data = datetime.strptime(data, formato)
      return data.strftime("%d/%m/%Y")
    except ValueError:
      continue
  return "erro"
def primeira_letra_maiuscula(string):
  if string is not None:
    if len(string) == 0:
      return string.upper()
    return string[0].upper()
def formatar_cep(cep):
    cep = ''.join(filter(str.isdigit, cep))
    if len(cep) != 8:
        return "erro"
    return f"{cep[:5]}-{cep[5:]}"
def remover_sinais(string):
  if string is not None:
    string = string.replace('.', '')
    string = string.replace('-', '')
    string = string.replace('_', '')
    string = string.replace('/', '')
  return string

# Conexão com banco de dados
mydb = mysql.connector.connect(
  host="localhost",
  user="root",
  password="",
  database="teste"
)
mycursor = mydb.cursor()

# Obtém data para busca de dados
dt_hoje = datetime.now()
dt_ontem = dt_hoje - timedelta(days=1)
data_busca = dt_ontem.strftime('%Y-%m-%d')

# Busca dados no banco
query = """SELECT 
        CASE WHEN rm.codigo_movimentacao = '01' THEN COALESCE(rm.nome_beneficiario_titular, bca.nome) WHEN rm.codigo_movimentacao = '02' THEN COALESCE(rm.nome_dependente, bca.nome) ELSE bca.nome END AS nome_beneficiario, 
        COALESCE (CASE WHEN rm.codigo_de_dependencia = 'Titular' THEN 'T' ELSE 'D' END, CASE WHEN bca.parentesco = 'Titular' THEN 'T' ELSE 'D' END) AS identificacao_beneficiario,
        rm.codigo_nome_plano,
        COALESCE (rm.data_nascimento, bca.dt_nascimento) AS data_nascimento, 
        '2' as tipo_endereco, 
        rm.logradouro AS endereco, 
        LTRIM(REPLACE(rm.numero_logradouro, '0', ' ')) AS numero_endereco, 
        rm.bairro_logradouro AS bairro_endereco, 
        rm.complemento_logradouro AS complemento_endereco, 
        rm.cep, 
        COALESCE (rm.uf, bca.estado) AS uf, 
        COALESCE (rm.municipio_cidade, bca.cidade) AS cidade,
        rm.numero_telefone, 
        COALESCE (rm.rg_documento, bca.rg) AS identidade, 
        rm.orgao_emissor_rg, 
        rm.estado_civil, 
        COALESCE (CASE WHEN rm.codigo_de_dependencia = 'Titular' THEN NULL ELSE rm.codigo_de_dependencia END, CASE WHEN bca.parentesco = 'Titular' THEN NULL ELSE bca.parentesco END) AS codigo_de_dependencia, 
        lf.credencial AS numero_carteirinha,
        CASE WHEN rm.codigo_movimentacao IN ('01', '02') THEN 'I' WHEN rm.codigo_movimentacao = '05' THEN 'E' END AS acao,
        COALESCE (rm.matricula, bca.chapa) AS matricula, 
        rm.cpf, 
        COALESCE (rm.sexo_genero, bca.sexo) AS sexo_genero, 
        COALESCE (rm.nome_mae, bca.mae) AS nome_mae,
        rm.pis, 
        rm.numero_nascido_vivo, 
        CASE WHEN clg.nome_coligada = 'MONTAGEM' THEN '915326' WHEN clg.nome_coligada = 'LSI SERVIÇOS' THEN '915073' WHEN clg.nome_coligada = 'LSI LOGISTICA' THEN '915328' WHEN clg.nome_coligada = 'FACILITIES' THEN '915074' ELSE NULL END AS cod_empresa, 
        CASE WHEN rm.codigo_movimentacao = '05' THEN DATE_FORMAT(NOW(), '%d/%m/%Y') ELSE NULL END AS hoje, 
        rm.codigo_movimentacao tipo_evento,
        rm.id 
        FROM registros_movimentacoes rm 
        LEFT JOIN base_colaborador_ativos bca ON rm.cpf = bca.cpf or rm.cpf_titular = bca.cpf
        LEFT JOIN linhas_faturas lf ON rm.cpf = lf.cpf 
        LEFT JOIN coligadas clg ON rm.local = clg.codigo 
        WHERE rm.data_movimentacao = '""" + data_busca + """';"""
mycursor.execute(query)
dtset = mycursor.fetchall()

# Cria uma cópia dos resultados obtidos formatando para o padrão do excel
novo_dtset = []
for item in dtset:
  novo_dtset.append({
    "A":item[0],
    "B":item[1],
    "C":item[2],
    "D":item[3],
    "E":item[4],
    "F":item[5],
    "G":item[6],
    "H":item[7],
    "I":item[8],
    "J":item[9],
    "K":item[10],
    "L":item[11],
    "M": "",
    "N":item[12],
    "O":item[13],
    "P":item[14],
    "Q":item[15],
    "R": "",
    "S":item[16],
    "T": item[17],
    "U":item[18],
    "V":item[19],
    "W":item[20],
    "X":item[21],
    "Y":item[22],
    "Z":item[23],
    "AA":item[24],
    "AB":item[25], 
    "AC":"",
    "AD":"",
    "AE":"N",
    "AF":"",
    "AG":"",
    "AH":"",
    "AI":"",
    "AJ":"",
    "AK":"",
    "AL":"",
    "AM":"",
    "AN":"",
    "AO":"",
    "AP":item[26],
    "AQ":item[27],
    "AR":item[28]
  })
dtset_excel = list(novo_dtset)

# Conversão de arquivo xls para xlsx
arq = r"C:\Users\giuli\Desktop\layout_web original_"
converte_xls_xlsx(arq+str(".xls"), arq+str(".xlsx"))

# Definindo caminho dos arquivos
arq_base = arq+str(".xlsx")
arq_gerado = r"C:\Users\giuli\Downloads\arqGerado.xlsx"

# Abrindo workbook e definindo linha para inicio do preenchimento
workbook = openpyxl.load_workbook(arq_base)
sheet = workbook.active
start_row = 2

# Definindo campos obrigatórios, semi obrigatórios e com regras
mandatory = {"A", "B", "C", "D", "F", "G", "J", "K", "U", "V", "X", "Y", "AB"}
semi_mandatory = {"S", "T", "W", "AA"}
with_rule = {"A", "D", "E", "H", "I", "J", "K", "O", "S", "U", "W", "X", "Y", "Z"}

# Loop para preenchimento dos dados linha a linha
for i, row in enumerate(novo_dtset):

  # Loop para validações de obrigatoriedade e regras
  erro = 0
  mensagem = []
  for j, (key, value) in enumerate(row.items()):

    # Verifica valores obrigatórios nulos
    if key in mandatory and value == None:
      match key:
        case "A":
          mensagem.append('"ERRO":"Sem nome do beneficiário"')
        case "B":
          mensagem.append('"ERRO":"Sem identificação do beneficiário"')
        case "C":
          mensagem.append('"ERRO":"Sem sigla do plano"')
        case "D":
          mensagem.append('"ERRO":"Sem data de nascimento"')
        case "F":
          mensagem.append('"ERRO":"Sem endereço"')
        case "G":
          mensagem.append('"ERRO":"Sem número do logradouro"')
        case "J":
          mensagem.append('"ERRO":"Sem CEP"')
        case "K":
          mensagem.append('"ERRO":"Sem estado/uf"')
        case "U":
          mensagem.append('"ERRO":"Sem ação"')
        case "V":
          mensagem.append('"ERRO":"Sem número de registro do funcionário"')
        case "X":
          mensagem.append('"ERRO":"Sem sexo"')
        case "Y":
          mensagem.append('"ERRO":"Sem nome da mãe"')
        case "AB":
          mensagem.append('"ERRO":"Código da empresa inválido"')
      erro = 1
      continue

    # Verifica valores semi obrigatórios nulos
    elif key in semi_mandatory and value == None :
      match key:
        case "S":
          if row["B"] is not None:
            if row["B"] == "D":
              erro = 1
              mensagem.append('"ERRO":"Sem parentesco definido"')
              continue
        case "T":
          if row["U"] is not None:
            if primeira_letra_maiuscula(value) != "I":
              erro = 1
              mensagem.append('"ERRO":"Sem número da carteirinha"')
              continue
        case "W":
          if row["D"] is not None:
            data_nascimento = datetime.strptime(row["D"], "%d/%m/%Y")
            diferenca = dt_hoje - data_nascimento
            idade = diferenca.days // 365
            if idade >= 18:
              erro = 1
              mensagem.append('"ERRO":"Sem CPF"')
              continue
        case "AA":
          if row["D"] is not None:
            data_nascimento = datetime.strptime(row["D"], "%d/%m/%Y")
            if data_nascimento.year >= 2010:
              erro = 1
              mensagem.append('"ERRO":"Sem número de declaração nascido vivo"')
              continue
    
    # Verifica regras de formatação dos dados
    if key in with_rule:
      match key:
        case "A": # Sem abreviação e limite de 70 caracteres
          if (verifica_abreviacao(value) == True):
            erro = 1
            mensagem.append('"ERRO":"Nome do beneficiário com abreviação"')
            continue
          else:
            dtset_excel[i]['A'] = limitar_string(value, 70)
        case "D": # Formato de data: dd/MM/yyyy
          if formatar_data(value) != "erro":
            dtset_excel[i]['D'] = formatar_data(value)
          else:
            erro = 1
            mensagem.append('"ERRO":"Nome do beneficiário com abreviação"')
            continue
        case "E": # Aceita apenas '1' ou '2'
          if primeira_letra_maiuscula(value) in {"P", "R", "1", "2"}:
            if primeira_letra_maiuscula(value) == "P":
              dtset_excel[i]['E'] = "1"
            elif primeira_letra_maiuscula(value) == "R":
              dtset_excel[i]['E'] = "2"
          else:
            erro = 1
            mensagem.append('"ERRO":"Tipo de endereço inválido"')
            continue
        case "H": # Limite de 20 caracteres
            dtset_excel[i]['H'] = limitar_string(value, 20)
        case "I": # Limite de 20 caracteres
            dtset_excel[i]['I'] = limitar_string(value, 20)
        case "J": # Formato de CEP: 00000-000
          cep = formatar_cep(value)
          if cep != "erro":
            dtset_excel[i]['J'] = cep
          else:
            erro = 1
            mensagem.append('"ERRO":"CEP inválido"')
            continue
        case "K":  # Limite de 2 caracteres
          dtset_excel[i]['K'] = limitar_string(value, 2)
        case "O":  # Remove sinais e valida 20 caracteres
          rg = remover_sinais(value)
          if rg is not None:
            if len(rg) > 20:
              erro = 1
              mensagem.append('"ERRO":"Identidade inválida"')
              continue
            else:
              dtset_excel[i]['O'] = rg
          else:
            dtset_excel[i]['O'] = rg
        case "S": # Limite de 10 caracteres
          dtset_excel[i]['S'] = limitar_string(value, 10)
        case "U": # Aceita apenas 'I', 'A', 'E', 'T', 'S' ou 'R'
          if primeira_letra_maiuscula(value) in {"I", "A", "E", "T", "S", "R"}:
            dtset_excel[i]['U'] = primeira_letra_maiuscula(value)
          else:
            erro = 1
            mensagem.append('"ERRO":"Ação inválida"')
            continue
        case "W": # Remove sinais e valida 11 caracteres
          cpf = remover_sinais(value)
          if cpf is not None:
            if len(cpf) > 11:
              erro = 1
              mensagem.append('"ERRO":"CPF inválido"')
              continue
            else:
              dtset_excel[i]['W'] = cpf
          else:
            dtset_excel[i]['W'] = cpf
        case "X": # Aceita apenas 'F' ou 'M'
          if primeira_letra_maiuscula(value) in {"F", "M"}:
            dtset_excel[i]['X'] = primeira_letra_maiuscula(value)
          else:
            erro = 1
            mensagem.append('"ERRO":"Sexo inválido"')
            continue
        case "Y": # Sem abreviação e limite de 70 caracteres
          if (verifica_abreviacao(value) == True):
            erro = 1
            mensagem.append('"ERRO":"Nome da mãe com abreviação"')
            continue
          else:
            dtset_excel[i]['Y'] = limitar_string(value, 70)
        case "Z": # Remove sinais e valida 11 caracteres
          pis = remover_sinais(value)
          if pis is not None:
            if len(pis) > 11:
              erro = 1
              mensagem.append('"ERRO":"PIS inválido"')
              continue
            else:
              dtset_excel[i]['Z'] = pis
          else:
            dtset_excel[i]['Z'] = pis
       
  # Escreve no excel e log sucesso ou log erro
  mensagem_string = ", ".join(mensagem)
  if erro == 0:
    sheet.append(dtset_excel[i])

    query = "INSERT INTO eventos_movimentacoes (dthr_evento, id_tipo_evento, id_resultado, created_at, id_registro_movimentacao) VALUES (now()," + row["AQ"] + ", 1, now(), " + str(row["AR"]) + ");"
    mycursor.execute(query)
    mydb.commit()
  else:
    query = "INSERT INTO eventos_movimentacoes (dthr_evento, mensagem, id_tipo_evento, id_resultado, created_at, id_registro_movimentacao) VALUES (now(), '" + mensagem_string + "', " + row["AQ"] + ", 2, now(), " + str(row["AR"]) + ");"
    mycursor.execute(query)
    mydb.commit()
      

# Salva o workbook modificado no novo arquivo
sheet.delete_cols(43, 2)
workbook.save(arq_gerado)