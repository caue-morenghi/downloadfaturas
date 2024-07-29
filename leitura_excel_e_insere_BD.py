from datetime import datetime
from openpyxl import load_workbook
import re
import pyodbc


#class InserirBD:

def InserirBD():
    #conexao com o BD
    hoje = datetime.now()
    dados_conexao = ("driver=MySQL ODBC 9.0 ANSI Driver;""server=localhost;""user=root;""database=grs;""password=;")
    cnx = pyodbc.connect(dados_conexao)
    cursor = cnx.cursor()

    #abrir o arquivo excel e definir a planilha
    wb = load_workbook('C:\\grs\\relatorios-selecionados (1)\\DIRF TRATAMENTO\\915069_915069-MANSERV INVES_072024.xlsx')
    planilha = wb["915069_915069-MANSERV INVES_072"]
    
    #lendo a planilha excel
    for i in range (1, planilha.max_row):
        chapa_do_funcionario = planilha.cell(row=i+1, column=1).value
        #cnpj_operadora = planilha.cell(row=i+1, column=2).value
        #nome_empresarial_operadora = planilha.cell(row=i+1, column=3).value
        registro_ans = planilha.cell(row=i+1, column=4).value
        cpf_titular = planilha.cell(row=i+1, column=5).value
        nome_titular = planilha.cell(row=i+1, column=6).value
        cpf_dependente = planilha.cell(row=i+1, column=7).value
        dt_nascto_dependente = planilha.cell(row=i+1, column=8).value
        nome_dependente = planilha.cell(row=i+1, column=9).value
        grau_relac_com_titular = planilha.cell(row=i+1, column=10).value
        ano_mes_ref = planilha.cell(row=i+1, column=11).value
        valor_pago = planilha.cell(row=i+1, column=12).value
        #tp_despesa = planilha.cell(row=i+1, column=13).value
        #numero_odonto_utilizador = planilha.cell(row=i+1, column=14).value

        #verificando se as linhas seguintes estão em branco
        if not (chapa_do_funcionario):
            break

        #fazendo um regex para tirar os espacos em branco depois do nome do dependente
        beneficiario = re.sub(r'\s+$', '', nome_dependente)
        titular = re.sub(r'\s+$', '', nome_titular)
        
        #pegando a coluna de mes e ano de referencia e separando o ano do mes que vem junto
        ano_mes_ref_str = str(ano_mes_ref)
        ano = ano_mes_ref_str[:4]
        mes = ano_mes_ref_str[4:]
        ano = int(ano)
        mes = int(mes)
        
        #transformando a data de nascimento que vem em datetime para date
        nascimento = dt_nascto_dependente.date()

        # verificacao para inserts na tabela
        if grau_relac_com_titular == 1:
            comando = f"""insert into linhas_faturas (matricula, plano, beneficiario, cobrado, cpf, nascimento, nome_titular, cpf_titular, parentesco) values ('{chapa_do_funcionario}', '{registro_ans}', '{beneficiario}', '{valor_pago}', '{cpf_dependente}', '{nascimento}', '{nome_titular}', '{cpf_titular}', 'CONJUGE')"""
            cursor.execute(comando)
            cursor.commit()

        elif grau_relac_com_titular == 3:
            comando = f"""insert into linhas_faturas (matricula, plano, beneficiario, cobrado, cpf, nascimento, nome_titular, cpf_titular, parentesco) values ('{chapa_do_funcionario}', '{registro_ans}', '{beneficiario}', '{valor_pago}', '{cpf_dependente}', '{nascimento}', '{nome_titular}', '{cpf_titular}', 'FILHO(A)')"""
            cursor.execute(comando)
            cursor.commit()
            
        else:
            comando = f"""insert into linhas_faturas (matricula, plano, beneficiario, cobrado, cpf, nascimento, nome_titular, cpf_titular, parentesco) values ('{chapa_do_funcionario}', '{registro_ans}', '{titular}', '{valor_pago}', '{cpf_titular}', '{nascimento}', '{nome_titular}', '{cpf_titular}', 'TITULAR')"""
            cursor.execute(comando)
            cursor.commit()

    comando = f"""insert into faturas (mes_base, ano_base) values ('{mes}', '{ano}')"""
    cursor.execute(comando)
    cursor.commit()


InserirBD()