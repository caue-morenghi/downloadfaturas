from datetime import datetime
from openpyxl import load_workbook
import re
import pyodbc
from pathlib import Path
import pandas as pd

def leitura_planilha():
    hoje = datetime.now()
    dados_conexao = ("driver=MySQL ODBC 9.0 ANSI Driver;""server=localhost;""user=root;""database=grs;""password=caue2005")
    cnx = pyodbc.connect(dados_conexao)
    cursor = cnx.cursor()

    caminho_planilha_csv = Path(r'C:\Users\Quaestum\Desktop\Cauê\robos\download faturas\915069_915069-MANSERV INVES_062024.csv')
    caminho_planilha_xlsx = Path(r'C:\Users\Quaestum\Desktop\Cauê\robos\download faturas\915069_915069-MANSERV INVES_062024.xlsx')
    
    df = pd.read_csv(caminho_planilha_csv, encoding='ISO-8859-1', decimal=',', sep=';', dtype=str)
    df.to_excel(caminho_planilha_xlsx, index=False, engine='openpyxl')

    if Path(caminho_planilha_xlsx).exists():
        wb = load_workbook(r'C:\Users\Quaestum\Desktop\Cauê\robos\download faturas\915069_915069-MANSERV INVES_062024.xlsx')
        planilha = wb["Sheet1"]

        for i in range (1, planilha.max_row):
            chapa_do_funcionario = planilha.cell(row=i+1, column=1).value
            #cnpj_operadora = planilha.cell(row=i+1, column=2).value
            #nome_empresarial_operadora = planilha.cell(row=i+1, column=3).value
            registro_ans = planilha.cell(row=i+1, column=4).value
            cpf_titular = planilha.cell(row=i+1, column=5).value
            nome_titular = planilha.cell(row=i+1, column=6).value
            cpf_dependente = planilha.cell(row=i+1, column=7).value
            dt_nascto_dependente = planilha.cell(row=i+1, column=8).value
            nome_dependente = planilha.cell(row=i+1, column=9).value
            grau_relac_com_titular = planilha.cell(row=i+1, column=10).value
            ano_mes_ref = planilha.cell(row=i+1, column=11).value
            valor_pago = planilha.cell(row=i+1, column=12).value
            #tp_despesa = planilha.cell(row=i+1, column=13).value
            #numero_odonto_utilizador = planilha.cell(row=i+1, column=14).value


leitura_planilha()